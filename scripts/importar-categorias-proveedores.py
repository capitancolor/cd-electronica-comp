import sqlite3
import requests
import openpyxl
import os
import sys

EXCEL_PATH = '/home/pablo/Escritorio/COSTOS 2_2.xlsx'
SQLITE_PATH = '/home/pablo/.config/com.cdinformatica.gestion/cd_electronica.db'
SUPABASE_URL = 'https://ewynsxiqohlnwbsemagb.supabase.co'
SUPABASE_KEY = 'sb_publishable_eI3nkVm2YkHQSZuCGVS29g_2YVVIBIx'

HEADERS = {
    'apikey': SUPABASE_KEY,
    'Authorization': f'Bearer {SUPABASE_KEY}',
    'Content-Type': 'application/json',
    'Prefer': 'return=representation'
}

def supabase_get(table):
    r = requests.get(f'{SUPABASE_URL}/rest/v1/{table}', headers={**HEADERS, 'Prefer': 'return=representation'}, params={'select': '*'})
    if r.status_code != 200:
        print(f'  Error GET {table}: {r.status_code} {r.text}')
        return []
    return r.json()

def supabase_upsert(table, data, on_conflict='id'):
    r = requests.post(
        f'{SUPABASE_URL}/rest/v1/{table}',
        headers={
            **HEADERS,
            'Prefer': f'resolution=merge-duplicates,return=representation',
            'on_conflict': on_conflict
        },
        json=data
    )
    if r.status_code not in (200, 201):
        print(f'  Error UPSERT {table}: {r.status_code} {r.text}')
        return None
    return r.json()

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f'❌ No se encuentra: {EXCEL_PATH}')
        sys.exit(1)

    if not os.path.exists(SQLITE_PATH):
        print(f'⚠️ No existe la DB local ({SQLITE_PATH}), se creará al conectar')

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]

    categorias = set()
    proveedores = set()

    for row in ws.iter_rows(min_row=4, values_only=True):
        cat = row[4]
        prov = row[7]
        if cat and str(cat).strip():
            categorias.add(str(cat).strip())
        if prov and str(prov).strip():
            proveedores.add(str(prov).strip())

    print(f'📊 Categorías únicas en Excel: {len(categorias)}')
    for c in sorted(categorias):
        print(f'    - {c}')
    print(f'\n📊 Proveedores únicos en Excel: {len(proveedores)}')
    for p in sorted(proveedores):
        print(f'    - {p}')

    # ── SUPABASE ──
    print('\n🔄 Sincronizando con Supabase...')

    cats_existing = {c['nombre'] for c in supabase_get('categorias')}
    provs_existing = {p['nombre'] for p in supabase_get('proveedores')}

    cats_to_insert = [{'nombre': c} for c in sorted(categorias) if c not in cats_existing]
    provs_to_insert = [{'nombre': p, 'activo': True} for p in sorted(proveedores) if p not in provs_existing]

    if cats_to_insert:
        print(f'  Insertando {len(cats_to_insert)} categorías nuevas...')
        result = supabase_upsert('categorias', cats_to_insert)
        if result:
            print(f'  ✅ Categorías insertadas/actualizadas')
    else:
        print(f'  ✅ Todas las categorías ya existen en Supabase')

    if provs_to_insert:
        print(f'  Insertando {len(provs_to_insert)} proveedores nuevos...')
        result = supabase_upsert('proveedores', provs_to_insert)
        if result:
            print(f'  ✅ Proveedores insertados/actualizados')
    else:
        print(f'  ✅ Todos los proveedores ya existen en Supabase')

    # ── SQLITE LOCAL ──
    print('\n🔄 Sincronizando con SQLite local...')

    conn = sqlite3.connect(SQLITE_PATH)
    cur = conn.cursor()

    # Crear tablas si no existen
    cur.execute('''
        CREATE TABLE IF NOT EXISTS categorias (
            id INTEGER PRIMARY KEY,
            nombre TEXT NOT NULL
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS proveedores (
            id INTEGER PRIMARY KEY,
            nombre TEXT NOT NULL,
            contacto TEXT,
            telefono TEXT,
            email TEXT,
            direccion TEXT,
            activo INTEGER DEFAULT 1
        )
    ''')

    # Obtener datos frescos de Supabase
    cats_supabase = supabase_get('categorias')
    provs_supabase = supabase_get('proveedores')

    cur.execute('DELETE FROM categorias')
    for c in cats_supabase:
        cur.execute('INSERT OR REPLACE INTO categorias (id, nombre) VALUES (?, ?)',
                    (c['id'], c['nombre']))
    conn.commit()
    print(f'  ✅ Categorías en SQLite: {len(cats_supabase)}')

    cur.execute('DELETE FROM proveedores')
    for p in provs_supabase:
        cur.execute(
            'INSERT OR REPLACE INTO proveedores (id, nombre, contacto, telefono, email, direccion, activo) VALUES (?, ?, ?, ?, ?, ?, ?)',
            (p['id'], p['nombre'], p.get('contacto'), p.get('telefono'),
             p.get('email'), p.get('direccion'), 1 if p.get('activo') else 0))
    conn.commit()
    print(f'  ✅ Proveedores en SQLite: {len(provs_supabase)}')

    conn.close()
    print('\n✅ Sincronización completada.')

if __name__ == '__main__':
    main()
